import os
import io
import csv
import json
import time
import random
import re
from dataclasses import dataclass
from datetime import datetime, timedelta, timezone
from typing import Dict, List, Optional, Tuple
from urllib.parse import quote

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# =========================
# CONFIG (V39)
# =========================
GRAPH_VERSION = os.getenv("GRAPH_VERSION", "v25.0").strip()
GRAPH_BASE = f"https://graph.facebook.com/{GRAPH_VERSION}"

STATE_FILE = os.getenv("STATE_FILE", "state.json").strip()
MAX_STATE_ITEMS = int(os.getenv("MAX_STATE_ITEMS", "9000"))

PAGE_ID = os.getenv("PAGE_ID", "").strip()
PAGE_ACCESS_TOKEN = os.getenv("PAGE_ACCESS_TOKEN", "").strip()
SHOPEE_CSV_URL = os.getenv("SHOPEE_CSV_URL", "").strip()

# Shopee Affiliate
AFFILIATE_ID = os.getenv("SHOPEE_AFFILIATE_ID", "15328100363").strip()
AFF_UTM_SOURCE = os.getenv("AFF_UTM_SOURCE", "facebook").strip()
AFF_TAG = os.getenv("AFF_TAG", "BENHomeElectrical").strip()

# Schedule (Bangkok)
TZ_BKK = timezone(timedelta(hours=7))
SLOTS_BKK = os.getenv("SLOTS_BKK", "12:00,18:30").split(",")
SLOTS_BKK = [s.strip() for s in SLOTS_BKK if s.strip()]

# Run behavior
POSTS_MAX_PER_RUN = int(os.getenv("POSTS_MAX_PER_RUN", "1"))  # ต่อ 1 run โพสต์ได้กี่โพสต์
FORCE_POST = os.getenv("FORCE_POST", "0").strip().lower() in ("1", "true", "yes")
FIRST_RUN_POST_1 = os.getenv("FIRST_RUN_POST_1", "1").strip().lower() in ("1", "true", "yes")

# Selection filters (ปรับให้ “ตรงเพจ”)
MIN_RATING = float(os.getenv("MIN_RATING", "4.7"))
MIN_DISCOUNT_PCT = float(os.getenv("MIN_DISCOUNT_PCT", "15"))
MIN_SOLD = int(os.getenv("MIN_SOLD", "50"))          # “เอาขายแล้วออกด้วย” = ขายต่ำกว่าตัดทิ้ง
PRICE_MIN = float(os.getenv("PRICE_MIN", "59"))
PRICE_MAX = float(os.getenv("PRICE_MAX", "4999"))
REPOST_AFTER_DAYS = int(os.getenv("REPOST_AFTER_DAYS", "21"))

# Media
POST_IMAGES_COUNT = int(os.getenv("POST_IMAGES_COUNT", "3"))

# CSV / XLSX streaming
STREAM_MAX_ROWS = int(os.getenv("STREAM_MAX_ROWS", "200000"))
TOPK_POOL = int(os.getenv("TOPK_POOL", "220"))

# Caption branding
BRAND = os.getenv("BRAND", "BEN Home & Electrical").strip()
HASHTAGS = os.getenv(
    "HASHTAGS",
    "#BENHomeElectrical #ของใช้ในบ้าน #อุปกรณ์ไฟฟ้า #เครื่องมือช่าง #งานช่าง #ซ่อมบ้าน #ShopeeAffiliate"
).strip()

# Category/Keyword targeting for BEN Home & Electrical
ALLOW_KEYWORDS = os.getenv(
    "ALLOW_KEYWORDS",
    r"(ไฟฟ้า|ปลั๊ก|พ่วง|เต้ารับ|สวิตช์|เบรกเกอร์|สายไฟ|หลอดไฟ|โคม|มิเตอร์|UPS|แบตเตอรี่|LiFePO4|อินเวอร์เตอร์|โซล่า|โซลาร์เซลล์|ชาร์จ|adapter|หัวชาร์จ|ปลั๊กแปลง|สายชาร์จ|type-?c|pd|qc|เครื่องมือ|สว่าน|ค้อน|ประแจ|คีม|ไขควง|ลูกบล็อก|ตลับเมตร|กาว|ซิลิโคน|เทป|งานช่าง|ซ่อมบ้าน|DIY|อุปกรณ์ช่าง|อุปกรณ์บ้าน)"
).strip()

BLOCK_KEYWORDS = os.getenv(
    "BLOCK_KEYWORDS",
    r"(เสื้อผ้า|กางเกง|ชุดนอน|หมอน|ผ้าห่ม|เครื่องสำอาง|ลิป|ครีม|สกินแคร์|อาหาร|ขนม|วิตามิน|อาหารเสริม|บุหรี่|แอลกอฮอล์|ย้อมผม|วิกผม|ตุ๊กตา|ของเล่น|แฟชั่น|บิกินี่|กระโปรง)"
).strip()

# Extra block by category text (กันหลุดหมวด)
BLOCK_CATEGORIES = os.getenv(
    "BLOCK_CATEGORIES",
    r"(Beauty|Health|Fashion|Women|Men|Shoes|Bags|Food|Groceries|Toys|Pets|Baby|Mom)"
).strip()

# =========================
# SAFETY CHECK
# =========================
def die(msg: str, code: int = 1):
    print("FATAL:", msg)
    raise SystemExit(code)

if not PAGE_ID:
    die("Missing env: PAGE_ID")
if not PAGE_ACCESS_TOKEN:
    die("Missing env: PAGE_ACCESS_TOKEN")
if not SHOPEE_CSV_URL:
    die("Missing env: SHOPEE_CSV_URL")

# =========================
# HTTP SESSION (retries)
# =========================
def make_session() -> requests.Session:
    s = requests.Session()
    retry = Retry(
        total=5,
        backoff_factor=0.6,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=("GET", "POST"),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    s.mount("https://", adapter)
    s.mount("http://", adapter)
    s.headers.update({"User-Agent": "fb-autopost-v39"})
    return s

SESS = make_session()

# =========================
# TIME / STATE
# =========================
def now_bkk() -> datetime:
    return datetime.now(TZ_BKK)

def parse_slot_today_bkk(now: datetime, hhmm: str) -> datetime:
    hh, mm = hhmm.split(":")
    return now.replace(hour=int(hh), minute=int(mm), second=0, microsecond=0)

def load_state() -> dict:
    base = {"used_urls": [], "posted_slots": {}, "posted_at": {}, "first_run_done": False}
    if not os.path.exists(STATE_FILE):
        return base
    try:
        with open(STATE_FILE, "r", encoding="utf-8") as f:
            s = json.load(f)
        if not isinstance(s, dict):
            return base
        s.setdefault("used_urls", [])
        s.setdefault("posted_slots", {})
        s.setdefault("posted_at", {})
        s.setdefault("first_run_done", True)
        if not isinstance(s["used_urls"], list):
            s["used_urls"] = []
        if not isinstance(s["posted_slots"], dict):
            s["posted_slots"] = {}
        if not isinstance(s["posted_at"], dict):
            s["posted_at"] = {}
        return s
    except Exception:
        return base

def save_state(state: dict) -> None:
    used = state.get("used_urls", [])
    if len(used) > MAX_STATE_ITEMS:
        state["used_urls"] = used[-MAX_STATE_ITEMS:]
    with open(STATE_FILE, "w", encoding="utf-8") as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

def mark_slot_posted(state: dict, now: datetime, slot_hhmm: str) -> None:
    key = now.strftime("%Y-%m-%d")
    state.setdefault("posted_slots", {})
    state["posted_slots"].setdefault(key, [])
    if slot_hhmm not in state["posted_slots"][key]:
        state["posted_slots"][key].append(slot_hhmm)

def due_slots_today(state: dict, now: datetime) -> List[str]:
    """ถ้าเลยเวลาแล้ว ยังไม่โพสต์ = ถือว่า due"""
    key = now.strftime("%Y-%m-%d")
    posted = set(state.get("posted_slots", {}).get(key, []))
    due = []
    for hhmm in SLOTS_BKK:
        t = parse_slot_today_bkk(now, hhmm)
        if now >= t and (hhmm not in posted):
            due.append(hhmm)
    return due

def pick_due_slot_or_none(state: dict, now: datetime) -> Optional[str]:
    due = due_slots_today(state, now)
    if due:
        # โพสต์ slot ที่ค้าง “เก่าสุด” ก่อน
        return due[0]
    return None

# =========================
# CSV/XLSX PARSER
# =========================
def fetch_bytes(url: str, timeout: int = 60) -> bytes:
    r = SESS.get(url, timeout=timeout)
    r.raise_for_status()
    return r.content

def iter_rows_from_csv_bytes(b: bytes) -> List[Dict[str, str]]:
    # ใช้ csv.DictReader (รองรับ header ไทย/อังกฤษ)
    text = b.decode("utf-8-sig", errors="replace")
    f = io.StringIO(text)
    reader = csv.DictReader(f)
    rows = []
    for i, row in enumerate(reader):
        if i >= STREAM_MAX_ROWS:
            break
        if not row:
            continue
        rows.append({k.strip(): (v.strip() if isinstance(v, str) else v) for k, v in row.items() if k})
    return rows

def iter_rows_from_xlsx_bytes(b: bytes) -> List[Dict[str, str]]:
    # อ่าน xlsx ด้วย openpyxl (ไม่ต้อง pandas)
    try:
        import openpyxl
    except Exception as e:
        die(f"Missing dependency openpyxl for XLSX: {e}")

    bio = io.BytesIO(b)
    wb = openpyxl.load_workbook(bio, data_only=True)
    ws = wb.active
    rows = []
    header = []
    for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
        if r_idx == 0:
            header = [str(x).strip() if x is not None else "" for x in row]
            continue
        if r_idx > STREAM_MAX_ROWS:
            break
        d = {}
        empty = True
        for c_idx, val in enumerate(row):
            key = header[c_idx] if c_idx < len(header) else ""
            if not key:
                continue
            sval = "" if val is None else str(val).strip()
            if sval:
                empty = False
            d[key] = sval
        if not empty and d:
            rows.append(d)
    return rows

def load_rows_from_url(url: str) -> List[Dict[str, str]]:
    b = fetch_bytes(url)
    u = url.lower()
    if u.endswith(".xlsx") or "format=xlsx" in u:
        return iter_rows_from_xlsx_bytes(b)
    return iter_rows_from_csv_bytes(b)

# =========================
# NORMALIZE ITEM
# =========================
def fnum(x: str, default: float = 0.0) -> float:
    try:
        if x is None:
            return default
        s = str(x).strip()
        if not s:
            return default
        s = re.sub(r"[^\d.\-]", "", s)
        return float(s)
    except Exception:
        return default

def inum(x: str, default: int = 0) -> int:
    try:
        if x is None:
            return default
        s = str(x).strip()
        if not s:
            return default
        s = re.sub(r"[^\d\-]", "", s)
        return int(float(s))
    except Exception:
        return default

def get_any(row: Dict[str, str], keys: List[str]) -> str:
    for k in keys:
        if k in row and str(row.get(k, "")).strip():
            return str(row.get(k, "")).strip()
    return ""

def collect_images(row: Dict[str, str]) -> List[str]:
    # รองรับหลายชื่อคอลัมน์
    candidates = []
    for k in list(row.keys()):
        lk = k.strip().lower()
        if lk in ("image_link", "image", "img", "imageurl", "image_url"):
            candidates.append(row.get(k, ""))
        if re.fullmatch(r"image_link_\d+", lk) or re.fullmatch(r"image_link\d+", lk):
            candidates.append(row.get(k, ""))
        if lk.startswith("image_link_") or lk.startswith("additional_image_link"):
            candidates.append(row.get(k, ""))
    # เพิ่มแบบระบุไว้ในตัวอย่าง
    for k in ["image_link", "image_link_2", "image_link_3", "image_link_4", "image_link_5", "image_link_6", "image_link_7", "image_link_8", "image_link_9", "image_link_10"]:
        if k in row:
            candidates.append(row.get(k, ""))

    out = []
    for u in candidates:
        u = (u or "").strip()
        if not u:
            continue
        if u.startswith("//"):
            u = "https:" + u
        if u.startswith("http://"):
            u = "https://" + u[len("http://"):]
        if u.startswith("https://") and u not in out:
            out.append(u)
    return out

def build_affiliate_url(product_link: str) -> str:
    # ใช้ shopee.ee redirect แบบที่คุณโพสต์ได้ผลแล้ว
    origin = quote(product_link, safe="")
    return (
        f"https://shopee.ee/an_redir?origin_link={origin}"
        f"&affiliate_id={quote(AFFILIATE_ID)}"
        f"&utm_source={quote(AFF_UTM_SOURCE)}"
        f"&afftag={quote(AFF_TAG)}"
    )

@dataclass
class Item:
    title: str
    product_link: str
    price: float
    sale_price: float
    discount_pct: float
    rating: float
    sold: int
    stock: int
    category_text: str
    description: str
    images: List[str]
    key: str  # unique id (prefer product_link)

def normalize_item(row: Dict[str, str]) -> Optional[Item]:
    title = get_any(row, ["title", "name", "product_name"])
    if not title:
        return None

    product_link = get_any(row, ["product_link", "product link", "product", "link", "url"])
    if not product_link:
        # บางไฟล์อาจมี shopid/itemid
        shopid = get_any(row, ["shopid"])
        itemid = get_any(row, ["itemid"])
        if shopid and itemid:
            product_link = f"https://shopee.co.th/product/{shopid}/{itemid}"
    if not product_link or "shopee" not in product_link:
        return None
    if product_link.startswith("http://"):
        product_link = "https://" + product_link[len("http://"):]
    if product_link.startswith("//"):
        product_link = "https:" + product_link

    price = fnum(get_any(row, ["price", "original_price", "model_price", "model_prices"]))
    sale_price = fnum(get_any(row, ["sale_price", "discount_price", "final_price"]))
    if sale_price <= 0:
        sale_price = price if price > 0 else 0

    discount_pct = fnum(get_any(row, ["discount_percentage", "discount_pct", "discount"]))
    rating = fnum(get_any(row, ["item_rating", "rating", "rate"]))
    sold = inum(get_any(row, ["item_sold", "sold", "historical_sold"]))
    stock = inum(get_any(row, ["stock", "quantity", "available_stock"]), default=0)

    cat1 = get_any(row, ["global_category1", "category1", "category"])
    cat2 = get_any(row, ["global_category2", "category2"])
    cat3 = get_any(row, ["global_category3", "category3"])
    category_text = " / ".join([c for c in [cat1, cat2, cat3] if c])

    description = get_any(row, ["description", "desc", "detail"])

    images = collect_images(row)
    if not images:
        return None

    key = product_link.strip()
    return Item(
        title=title.strip(),
        product_link=product_link.strip(),
        price=price,
        sale_price=sale_price,
        discount_pct=discount_pct,
        rating=rating,
        sold=sold,
        stock=stock,
        category_text=category_text,
        description=description,
        images=images,
        key=key,
    )

# =========================
# FILTERS (ให้ตรงเพจ)
# =========================
ALLOW_RE = re.compile(ALLOW_KEYWORDS, re.IGNORECASE)
BLOCK_RE = re.compile(BLOCK_KEYWORDS, re.IGNORECASE)
BLOCK_CAT_RE = re.compile(BLOCK_CATEGORIES, re.IGNORECASE)

def is_relevant_to_page(it: Item) -> bool:
    blob = f"{it.title}\n{it.category_text}\n{it.description}"
    if BLOCK_RE.search(blob):
        return False
    if it.category_text and BLOCK_CAT_RE.search(it.category_text):
        return False
    return bool(ALLOW_RE.search(blob))

def pass_metrics(it: Item) -> bool:
    if it.stock <= 0:
        return False
    if it.sale_price < PRICE_MIN or it.sale_price > PRICE_MAX:
        return False
    if it.rating < MIN_RATING:
        return False
    if it.discount_pct < MIN_DISCOUNT_PCT:
        return False
    if it.sold < MIN_SOLD:
        return False
    return True

def score_item(it: Item) -> float:
    # คะแนนสูง = ดี: ส่วนลด + เรตติ้ง + sold + ตรงหมวด
    # ทำให้ "ของไฟฟ้า/เครื่องมือ" มักขึ้นก่อน
    base = 0.0
    base += min(it.discount_pct, 90) * 1.5
    base += min(it.rating, 5.0) * 12.0
    base += min(it.sold, 5000) ** 0.5 * 4.0
    if is_relevant_to_page(it):
        base += 30.0
    # ราคากลาง ๆ ดูดีกว่า (ไม่ให้ของแพงเกินชนะง่าย)
    if 200 <= it.sale_price <= 1500:
        base += 10.0
    return base

# =========================
# GRAPH API HELPERS
# =========================
def graph_post(path: str, data: dict, timeout: int = 60) -> dict:
    url = GRAPH_BASE + path
    payload = dict(data)
    payload["access_token"] = PAGE_ACCESS_TOKEN
    r = SESS.post(url, data=payload, timeout=timeout)
    try:
        js = r.json()
    except Exception:
        js = {"error": {"message": r.text}}
    if r.status_code >= 400 or ("error" in js):
        raise RuntimeError(f"Graph API error: {js}")
    return js

def upload_unpublished_photo(page_id: str, image_url: str) -> str:
    js = graph_post(
        f"/{page_id}/photos",
        data={"url": image_url, "published": "false"},
        timeout=90,
    )
    return js["id"]

def create_feed_post_with_media(page_id: str, message: str, media_fbids: List[str]) -> str:
    data = {"message": message}
    for i, mid in enumerate(media_fbids):
        data[f"attached_media[{i}]"] = json.dumps({"media_fbid": mid})
    js = graph_post(f"/{page_id}/feed", data=data, timeout=90)
    return js["id"]

# =========================
# CAPTION
# =========================
def format_baht(x: float) -> str:
    try:
        return f"{int(round(x)):,}"
    except Exception:
        return str(x)

def pick_bullets(it: Item) -> List[str]:
    # บูลเล็ตให้เข้าธีมเพจ
    bullets = []
    t = it.title.lower()
    blob = (it.title + " " + it.description).lower()

    if re.search(r"(ปลั๊ก|พ่วง|เต้ารับ|สวิตช์|สายไฟ|เบรกเกอร์|หลอดไฟ|โคม)", blob):
        bullets.append("เหมาะกับงานไฟฟ้าในบ้าน/ร้าน")
    if re.search(r"(สว่าน|คีม|ไขควง|ประแจ|ค้อน|ลูกบล็อก|เครื่องมือ)", blob):
        bullets.append("งานช่าง/DIY ใช้ได้จริง")
    if re.search(r"(ชาร์จ|adapter|หัวชาร์จ|type-?c|pd|qc|สายชาร์จ)", blob):
        bullets.append("ชาร์จไว พกง่าย ใช้คุ้ม")

    bullets.append("ดูรูป/รีวิวก่อนซื้อได้ ✅")
    # จำกัด 2–3 บรรทัด
    return bullets[:3]

def build_caption(it: Item) -> str:
    aff_url = build_affiliate_url(it.product_link)

    promo = f"{format_baht(it.sale_price)} บาท"
    normal = f"{format_baht(it.price)}" if it.price > 0 else ""
    disc = f"{int(round(it.discount_pct))}%"
    rating = f"{it.rating:.1f}/5" if it.rating > 0 else "-"
    sold = f"{it.sold:,}"

    bullets = pick_bullets(it)
    bullet_txt = "\n".join([f"✅ {b}" for b in bullets])

    lines = [
        f"🏠⚡ {BRAND}",
        "🔥 คัดของเด็ดสายบ้าน/ช่าง — โปรแรง รีวิวดียอดขายจริง",
        "",
        f"🛒 {it.title}",
        "",
        f"💸 โปรวันนี้: {promo}" + (f" (ปกติ {normal} | ลด {disc})" if normal else f" (ลด {disc})"),
        f"⭐ เรตติ้ง: {rating}",
        f"📦 ขายแล้ว: {sold} ชิ้น",
        "",
        bullet_txt,
        "",
        "👉 กดลิงก์ดูโปร/โค้ดส่วนลดได้ที่นี่",
        aff_url,
        "",
        HASHTAGS,
    ]
    return "\n".join([x for x in lines if x is not None])

# =========================
# PICK ITEM (NO DUP / REPOST WINDOW)
# =========================
def is_recently_posted(state: dict, key: str, now: datetime) -> bool:
    posted_at = state.get("posted_at", {}).get(key)
    if not posted_at:
        return False
    try:
        dt = datetime.fromisoformat(posted_at)
    except Exception:
        return False
    return (now - dt) < timedelta(days=REPOST_AFTER_DAYS)

def mark_posted(state: dict, key: str, now: datetime) -> None:
    state.setdefault("posted_at", {})
    state["posted_at"][key] = now.isoformat()
    state.setdefault("used_urls", [])
    state["used_urls"].append(key)

def choose_best_item(rows: List[Dict[str, str]], state: dict, now: datetime) -> Optional[Item]:
    pool: List[Item] = []

    for row in rows:
        it = normalize_item(row)
        if not it:
            continue
        if is_recently_posted(state, it.key, now):
            continue
        if not is_relevant_to_page(it):
            continue
        if not pass_metrics(it):
            continue
        pool.append(it)
        if len(pool) >= TOPK_POOL:
            break

    if not pool:
        return None

    # สุ่มใน top ที่คะแนนสูง เพื่อลดการโพสต์ซ้ำแนวเดิม
    pool.sort(key=score_item, reverse=True)
    top = pool[: min(25, len(pool))]
    weights = [max(1.0, score_item(x)) for x in top]
    # random.choices ต้องการ weights
    pick = random.choices(top, weights=weights, k=1)[0]
    return pick

# =========================
# MAIN
# =========================
def main():
    now = now_bkk()
    state = load_state()

    # 1) decide if should post
    due_slot = pick_due_slot_or_none(state, now)

    if not state.get("first_run_done", False) and FIRST_RUN_POST_1:
        should_post = True
        slot_to_mark = None
        print("INFO: First run detected -> post 1 immediately")
    else:
        if FORCE_POST:
            should_post = True
            slot_to_mark = due_slot or "FORCE"
            print("INFO: FORCE_POST enabled")
        else:
            should_post = due_slot is not None
            slot_to_mark = due_slot
            if not should_post:
                print("INFO: No due slot (already posted or not yet time)")
                return

    # 2) load rows
    print("INFO: Loading feed file:", SHOPEE_CSV_URL)
    rows = load_rows_from_url(SHOPEE_CSV_URL)
    print("INFO: Rows loaded:", len(rows))

    # 3) pick best item
    it = choose_best_item(rows, state, now)
    if not it:
        print("INFO: No eligible item found (filters too strict or out of stock)")
        # ยัง mark slot ได้ไหม? ไม่ mark เพื่อไม่ให้เสีย slot
        return

    # 4) choose images
    imgs = it.images[: max(1, POST_IMAGES_COUNT)]
    if not imgs:
        print("INFO: No images -> skip")
        return

    # 5) up
